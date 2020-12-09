# import web driver
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook
import os 

login_username = ""
login_password = ""

cwd = os.getcwd()
webdriver_path = cwd + "/deps/chromedriver"

#List of Companies
companies = ['m4u', "stone-pagamentos", "radix-engenharia-e-software", "b2w---companhia-global-do-varejo", "picpay", "mandic"]
#List of Keywords
keywords = ["DevOps"]
# Max Number of Pages to search for keyword
max_pages = 5 


def create_driver():
    # specifies the path to the chromedriver.exe
    return webdriver.Chrome(webdriver_path)

def login(driver):
    #will navigate to a page given by the URL address
    driver.get('https://www.linkedin.com')

    username = driver.find_element_by_xpath('//*[@id="session_key"]')
    username.send_keys(login_username)
    password = driver.find_element_by_xpath('//*[@id="session_password"]')
    password.send_keys(login_password)

    sign_in_button = driver.find_element_by_xpath('/html/body/main/section[1]/div[2]/form/button')
    sign_in_button.click()


def get_company_page(driver, company):
    company_url = "https://www.linkedin.com/company/{}/".format(company)
    print("Search for company url: {}".format(company_url))
    driver.get(company_url)
    driver.implicitly_wait(5)

def search_people(driver):
    people_link = driver.find_element_by_link_text('People')
    people_link.click()
    driver.implicitly_wait(5)

def nav_home(driver):
    people_link = driver.find_element_by_link_text('Home')
    people_link.click()
    driver.implicitly_wait(5)    
    
def query_keyword(driver, keyword):
    people_search = driver.find_element_by_xpath('//*[@id="people-search-keywords"]')
    people_search.send_keys(keyword)
    people_search.send_keys(Keys.RETURN)
    driver.implicitly_wait(5)


def scroll_top(driver):
    driver.execute_script("window.scrollTo(0, 200);")
    driver.implicitly_wait(5)

def scroll_until_end_of_list(driver):
    SCROLL_PAUSE_TIME = 2.0
    # Get scroll height
    last_height = driver.execute_script("return document.body.scrollHeight")
    page = 1 
    while page < max_pages:
        # Scroll down to bottom
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        # Wait to load page
        time.sleep(SCROLL_PAUSE_TIME)
        # Calculate new scroll height and compare with last scroll height
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height
        page += 1

def parse_people(driver):
    content = driver.page_source
    soup = BeautifulSoup(content, 'html.parser')
    people_containers = soup.find_all(class_="org-people-profile-card")
    people = []
    for container in people_containers:
        profile_element = container.find(class_="org-people-profile-card__profile-title")
        if profile_element is None:
            continue
        person_name = profile_element.get_text()
        links = container.find_all('a', href=True)
        person_link = "https://www.linkedin.com"
        for link in links:
            href = link['href']
            if "/in/" in href:
                person_link += href
                break
        line = {"name": person_name.strip(), "link": person_link}
        people.append(line)
    return people    


def create_mappings(driver, companies, keywords):
    mappings = {}
    for company in companies:
        mappings[company] = {}
        get_company_page(driver, company)
        for keyword in keywords:
            search_people(driver)
            query_keyword(driver, keyword)
            query_keyword(driver, "Brazil")
            scroll_until_end_of_list(driver)
            people = parse_people(driver)
            mappings[company][keyword] = people
            scroll_top(driver)
            nav_home(driver)

    return mappings

def create_header(ws):
    columns = ["company", "keyword", "employee", "profile_link"]
    colIdx = 1
    for col in columns:
        ws.cell(row=1, column=colIdx, value=col)
        colIdx += 1

# 
# mappings format
# {
#   company: {
#       keyword: [
#           name: "",
#           link: ""
#       ]
#   }
# }
#     
def write_to_excel(mappings):
    wb = Workbook()
    ws = wb.active
    create_header(ws)
    rowIdx = 2
    company_col_idx, keyword_col_idx, employee_col_idx, profile_link_col_idx = 1, 2, 3, 4
    companies = mappings.keys()
    for company in companies:
        keywords = mappings[company].keys()
        for keyword in keywords:
            employees = mappings[company][keyword]
            for employee in employees:
                ws.cell(row=rowIdx, column=company_col_idx, value=company)
                ws.cell(row=rowIdx, column=keyword_col_idx, value=keyword)
                ws.cell(row=rowIdx, column=employee_col_idx, value=employee["name"])
                ws.cell(row=rowIdx, column=profile_link_col_idx, value=employee["link"])
                rowIdx += 1    
    
    wb.save('leads.xlsx')        

driver = create_driver()
login(driver)
mappings = create_mappings(driver, companies, keywords)
write_to_excel(mappings)
driver.close()
